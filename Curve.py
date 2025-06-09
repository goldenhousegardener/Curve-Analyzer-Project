import cv2
import numpy as np
import math
import os
import tkinter as tk
from tkinter import filedialog, scrolledtext
from PIL import Image, ImageDraw, ImageFont
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import pytesseract

# Dictionary to store axio values per patient
patient_axio_values = {}

def log_message(message):
    """Logs messages to the UI text box and console."""
    log_text.insert(tk.END, message + "\n")
    log_text.see(tk.END)
    print(message)

def preprocess_image(image_path):
    """Loads and processes the image (grayscale, blur, edge detection)."""
    image = cv2.imread(image_path)
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    edges = cv2.Canny(blurred, 50, 150)
    return image, gray, edges

def is_white_background(gray_image, x, y, w, h, threshold=200):
    """Checks if a region has a white background based on intensity threshold."""
    roi = gray_image[y:y+h, x:x+w]
    mean_intensity = np.mean(roi)
    return mean_intensity > threshold

def merge_bounding_boxes(boxes):
    """Merges multiple bounding boxes into one."""
    if not boxes:
        return None
    x_min = min(box[0] for box in boxes)
    y_min = min(box[1] for box in boxes)
    x_max = max(box[0] + box[2] for box in boxes)
    y_max = max(box[1] + box[3] for box in boxes)
    return (x_min, y_min, x_max - x_min, y_max - y_min)

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os

def save_blue_dots_to_csv(patient_name, all_sections_data, result_folder):
    """Saves blue dot angles into result.xlsx, updating existing patient row or appending if new."""
    
    # Define the fixed header columns as per the expected structure
    header_columns = ["Patient"]
    record_types = ["AD", "AI", "BD", "BI"]
    time_points = ["T1", "T2", "T3", "T4"]
    for record in record_types:
        for time in time_points:
            header_columns.append(f"{record}_{time}")
    
    # Add the new axio columns
    header_columns.extend(["axio dei", "axio izq"])
    
    # Define file path
    excel_filename = os.path.join(result_folder, "result.xlsx")
    
    # For debugging: print patient_name and all_sections_data keys
    print(f"Saving data for Patient: {patient_name}")
    # print(f"All sections data keys: {list(all_sections_data.keys())}")
    
    if os.path.exists(excel_filename):
        wb = openpyxl.load_workbook(excel_filename)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Blue Dot Angles"
        ws.append(header_columns)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = header_font
            cell.fill = header_fill
    
    # Prepare row data with empty values for all columns except Patient
    row_data = [""] * len(header_columns)
    row_data[0] = patient_name  # Patient column
    
    # Set axio values if available for this patient
    # Try both the full patient name and just the folder name
    patient_values = patient_axio_values.get(patient_name, {})
    if not patient_values:
        # Try getting values using just the folder name (last part of patient name)
        folder_name = patient_name.split()[-1] if patient_name else ""
        patient_values = patient_axio_values.get(folder_name.upper(), {})
    
    row_data[-2] = patient_values.get('dei')  # axio dei
    row_data[-1] = patient_values.get('izq')  # axio izq
    print(f"Looking up axio values for patient '{patient_name}' (folder: {folder_name if 'folder_name' in locals() else 'N/A'})")
    print(f"Found values: dei={patient_values.get('dei')}, izq={patient_values.get('izq')}")
    
    # Fill the row_data with angles for all record types and time points
    # all_sections_data is expected to be a dict with keys as record_types and values as dict of section_index to angles list
    for record_index, record_type in enumerate(record_types):
        sections_data = all_sections_data.get(record_type, {})
        # Flatten angles from all sections for this record type
        angles_flat = []
        for section_index in sorted(sections_data.keys()):
            angles_flat.extend(sections_data[section_index])
        # Fill up to 4 angles for this record type
        start_index = 1 + record_index * len(time_points)
        for i in range(min(len(angles_flat), len(time_points))):
            row_data[start_index + i] = angles_flat[i]
        # print(f"Record type {record_type} angles_flat: {angles_flat}"
    
    # Check if patient already exists in sheet
    patient_row = None
    for row in ws.iter_rows(min_row=2, max_col=1):
        cell = row[0]
        if cell.value == patient_name:
            patient_row = cell.row
            break
    
    if patient_row:
        # Update existing row
        for col_index, value in enumerate(row_data, start=1):
            ws.cell(row=patient_row, column=col_index, value=value)
    else:
        # Append new row
        ws.append(row_data)
    
    # Auto-adjust column width
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    # Save the file
    wb.save(excel_filename)
    print(f"Data saved to {excel_filename}")
    return excel_filename

def detect_graph_area(edges, gray_image, image_width):
    """Finds the graph area on the rightmost 1/4 of the image."""
    contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    graph_areas = []
    min_x_threshold = image_width * 0.7  # Rightmost 1/4

    for cnt in contours:
        x, y, w, h = cv2.boundingRect(cnt)
        if w > 10 and h > 10 and x > min_x_threshold and is_white_background(gray_image, x, y, w, h):
            graph_areas.append((x, y, w, h))

    return merge_bounding_boxes(graph_areas)

def detect_blue_points(image, roi):
    """Detects isolated blue points while filtering out connected regions (lines)."""
    if roi is None:
        return []

    x, y, w, h = roi
    roi_image = image[y:y+h, x:x+w]

    # Convert to HSV
    hsv = cv2.cvtColor(roi_image, cv2.COLOR_BGR2HSV)
    lower_blue = np.array([100, 100, 100])
    upper_blue = np.array([140, 255, 255])
    mask = cv2.inRange(hsv, lower_blue, upper_blue)

    # Find contours in the blue mask
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    blue_dots = []
    for cnt in contours:
        area = cv2.contourArea(cnt)
        if 5 < area < 100:  # Filter out noise (too small) and lines (too large)
            M = cv2.moments(cnt)
            if M["m00"] != 0:
                cx = int(M["m10"] / M["m00"])  # Compute centroid
                cy = int(M["m01"] / M["m00"])
                blue_dots.append((cx + x, cy + y))  # Adjust for ROI offset
    # if len(blue_dots) > 1:
    #     one_dot = []
    #     one_dot.append(blue_dots[0])
    #     return one_dot
    return blue_dots

def split_graph_area(graph_area, gray_image):
    """Automatically splits the graph area based on white and black regions."""
    if graph_area is None:
        return []
    
    x, y, w, h = graph_area
    roi_gray = gray_image[y:y+h, x:x+w]
    
    # Compute horizontal projection (sum of pixel intensities along rows)
    projection = np.sum(roi_gray, axis=1)
    
    # Normalize projection values to range [0, 255]
    projection = (projection - np.min(projection)) / (np.max(projection) - np.min(projection)) * 255
    
    # Identify significant gaps (peaks in the projection)
    gaps = np.where(projection > 200)[0]  # Threshold for white space
    
    # Identify split points by detecting significant gaps between black regions
    split_indices = []
    prev_gap = gaps[0] if len(gaps) > 0 else 0
    for gap in gaps:
        if gap - prev_gap > 20:  # Ensure a minimum distance between splits
            split_indices.append(gap)
        prev_gap = gap
    
    # Create section bounding boxes based on detected splits
    sections = []
    start_y = y
    for split_y in split_indices:
        sections.append((x, start_y, w, split_y - start_y))
        start_y = split_y + y
    
    # Add the last section if needed
    if start_y < y + h:
        sections.append((x, start_y, w, (y + h) - start_y))
    
    return sections

def detect_red_curve(image, roi):
    """Detects the red curve in a given region of interest (ROI)."""
    if roi is None:
        return []
    
    x, y, w, h = roi
    roi_image = image[y:y+h, x:x+w]
    
    # Convert to HSV color space
    hsv = cv2.cvtColor(roi_image, cv2.COLOR_BGR2HSV)
    
    # Define red color range (considering both lower and upper red hues)
    lower_red1 = np.array([0, 120, 70])
    upper_red1 = np.array([10, 255, 255])
    lower_red2 = np.array([170, 120, 70])
    upper_red2 = np.array([180, 255, 255])
    
    mask1 = cv2.inRange(hsv, lower_red1, upper_red1)
    mask2 = cv2.inRange(hsv, lower_red2, upper_red2)
    mask = cv2.bitwise_or(mask1, mask2)
    
    # Find contours of the red curve
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    red_curve_points = []
    for cnt in contours:
        for point in cnt:
            red_curve_points.append((point[0][0] + x, point[0][1] + y))

    return red_curve_points

def detect_axes(image, roi):
    """Finds and marks perpendicular x and y axes using thick black lines."""
    if roi is None:
        return None, None, None

    x, y, w, h = roi
    roi_gray = cv2.cvtColor(image[y:y+h, x:x+w], cv2.COLOR_BGR2GRAY)

    # Threshold to highlight black lines
    _, binary = cv2.threshold(roi_gray, 100, 255, cv2.THRESH_BINARY_INV)

    # Detect lines
    lines = cv2.HoughLinesP(binary, 1, np.pi / 180, 80, minLineLength=30, maxLineGap=15)

    if lines is None:
        return None, None, None

    vertical_lines = []
    horizontal_lines = []

    for line in lines:
        x1, y1, x2, y2 = line[0]
        if abs(x1 - x2) < 20:  # Vertical line
            vertical_lines.append(((x1 + x, y1 + y), (x2 + x, y2 + y)))
        elif abs(y1 - y2) < 20:  # Horizontal line
            horizontal_lines.append(((x1 + x, y1 + y), (x2 + x, y2 + y)))

    if not vertical_lines or not horizontal_lines:
        return None, None, None

    v_line = vertical_lines[0]
    h_line = horizontal_lines[0]

    cv2.line(image, v_line[0], v_line[1], (0, 0, 0), 1)  # Black vertical
    cv2.line(image, h_line[0], h_line[1], (0, 0, 0), 1)  # Black horizontal

    intersection = (v_line[0][0], h_line[0][1])
    cv2.circle(image, intersection, 5, (0, 0, 255), -1)  # Red point

    cv2.putText(image, "(0, 0)", (intersection[0] + 10, intersection[1] - 10),
                cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 0), 1, cv2.LINE_AA)
    return vertical_lines[0], horizontal_lines[0], (vertical_lines[0][0][0], horizontal_lines[0][0][1])

def process_image(input_path, save_path, all_sections_data=None):
    """Processes a single image and returns patient name and all_sections_data."""
    image, gray, edges = preprocess_image(input_path)
    image_height, image_width = image.shape[:2]

    graph_area = detect_graph_area(edges, gray, image_width)
    graph_sections = split_graph_area(graph_area, gray)

    # Parse patient name and record type from filename
    filename = os.path.splitext(os.path.basename(input_path))[0]
    parts = filename.split()
    if len(parts) >= 2:
        patient_name = " ".join(parts[:-1]).upper()
        record_type = parts[-1].upper()
    else:
        patient_name = filename.upper()
        record_type = ""
    # print("~~~~~~~~~~", record_type)
    # Check if record_type is known, else warn but keep original
    valid_record_types = {"AD", "AI", "BD", "BI"}
    if record_type not in valid_record_types:
        print(f"Warning: record_type '{record_type}' not recognized.")
        # Keep record_type as is

    # Prepare a dictionary to hold all record types' sections data if not provided
    if all_sections_data is None:
        all_sections_data = {
            "AD": {},
            "AI": {},
            "BD": {},
            "BI": {}
        }

    # Define mapping of section index to record type for AI and BI images
    section_to_record_type_map = {
        "AI": {0: "AD", 1: "AI"},
        "BI": {0: "BD", 1: "BI"}
    }
    section_index_const = -1
    for section_index, roi in enumerate(graph_sections):
        v_line, h_line, intersection = detect_axes(image, roi)
        blue_dots = detect_blue_points(image, roi)
        red_curve = detect_red_curve(image, roi)
        section_index_const = section_index_const + 1
        # print("section_index", section_index)
        if intersection and blue_dots and red_curve:
            # Calculate angles for this section
            angles_list = []
            blue_dots.sort(key=lambda p: p[0])
            for (x_blue, y_blue) in blue_dots:
                cv2.circle(image, (x_blue, y_blue), 5, (255, 0, 0), -1)
                cv2.line(image, intersection, (x_blue, y_blue), (255, 0, 0), 2)
                cv2.line(image, (intersection[0], y_blue), (x_blue, y_blue), (0, 0, 0), 1)

                x_rel = x_blue - intersection[0]
                y_rel = intersection[1] - y_blue
                angle = math.degrees(math.atan2(y_rel, x_rel))
                
                if abs(angle) > 90:
                    angle = 180 - abs(angle)
                
                image_pil = Image.fromarray(cv2.cvtColor(image, cv2.COLOR_BGR2RGB))
                draw = ImageDraw.Draw(image_pil)
                font = ImageFont.truetype("arial.ttf", 15)

                draw.text((x_blue - 40, y_blue + 10), f"{math.fabs(angle):.2f}°", font=font, fill=(0, 0, 0))
                draw.text((x_blue - 40, y_blue + 28), f"({x_rel}, {y_rel})", font=ImageFont.truetype("arial.ttf", 12), fill=(0, 0, 0))
                image = cv2.cvtColor(np.array(image_pil), cv2.COLOR_RGB2BGR)

                # Get four equal divisions of the blue line
                x1, y1 = intersection
                x2, y2 = (x_blue, y_blue)

                closest_point_arr = []
                
                for i in range(1, 4):  # Divide into 4 equal parts
                    px = int(x1 + (x2 - x1) * (i / 4))
                    py = int(y1 + (y2 - y1) * (i / 4))

                    # Find the closest point on the red curve
                    min_dist = float('inf')
                    closest_point = None
                    for point in red_curve:
                        cx, cy = point
                        dist = np.linalg.norm(np.array((px, py)) - np.array((cx, cy)))
                        if dist < min_dist:
                            min_dist = dist
                            closest_point = (cx, cy)

                    if closest_point:
                        cv2.circle(image, closest_point, 5, (0, 165, 255), -1)  # Orange closest points
                        cv2.line(image, (intersection[0], closest_point[1]), closest_point, (0, 0, 0), 1)
                        closest_point_arr.append(closest_point)

                x_vir, y_vir = closest_point_arr[0]
                x_rel1 = x_vir - intersection[0]
                y_rel1 = intersection[1] - y_vir
                angle_1 = math.degrees(math.atan2(y_rel1, x_rel1))
                if math.fabs(angle_1) > 90:
                    angle_1 = 180 - math.fabs(angle_1)
                angles_list.append(abs(angle_1))
                cv2.putText(image, f"{math.fabs(angle_1):.2f}", (x_vir + 10, y_vir),
                cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 0), 1, cv2.LINE_AA)

                cv2.line(image, closest_point_arr[0], intersection, (0, 0, 0), 1)  # Draw orange line
                cv2.line(image, closest_point_arr[-1], (x_blue, y_blue), (0, 0, 0), 1)  # Draw orange line

                if len(closest_point_arr) > 1:
                    for i in range(len(closest_point_arr) - 1):
                        x_vir_i, y_vir_i = closest_point_arr[i + 1]
                        y_vir_i = closest_point_arr[i + 1][1]
                        x_vir_2, y_vir_2 = closest_point_arr[i]
                        x_rel1 = x_vir_i - x_vir_2
                        y_rel1 = y_vir_2 - y_vir_i
                        angle_1_i = math.degrees(math.atan2(y_rel1, x_rel1))
                        if math.fabs(angle_1_i) > 90:
                            angle_1_i = 180 - math.fabs(angle_1_i)
                        cv2.putText(image, f"{math.fabs(angle_1_i):.2f}", (x_vir_i + 10, y_vir_i),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 0), 1, cv2.LINE_AA)
                        cv2.line(image, closest_point_arr[i], closest_point_arr[i + 1], (0, 0, 0), 1)  # Draw orange line
                        angles_list.append(abs(angle_1_i))

                x, y = closest_point_arr[-1]
                x_rel1 = x_blue - x
                y_rel1 = y - y_blue
                angle_1 = math.degrees(math.atan2(y_rel1, x_rel1))
                if math.fabs(angle_1) > 90:
                    angle_1 = 180 - math.fabs(angle_1)
                angles_list.append(abs(angle_1))
                cv2.putText(image, f"{math.fabs(angle_1):.2f}", (x_blue + 10, y_blue),
                cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 0), 1, cv2.LINE_AA)

                angles_list.append(abs(angle))

            # Determine the actual record type for this section based on mapping
            # print("!!!!!!!!!", record_type, section_index_const, angles_list)
            actual_record_type = record_type
            if record_type in section_to_record_type_map:
                actual_record_type = section_to_record_type_map[record_type].get(section_index_const, record_type)

            # Assign angles_list to the corresponding actual_record_type in all_sections_data
            # print("Assigning angles_list to record_type:", actual_record_type)
            if actual_record_type in all_sections_data:
                all_sections_data[actual_record_type][section_index_const] = angles_list
                # print(f"Assigned angles_list for {actual_record_type} section {section_index_const}: {angles_list}")
            else:
                print(f"Warning: record_type '{actual_record_type}' not in all_sections_data keys. Skipping assignment.")
        else:
            section_index_const = section_index_const - 1
    # Save analyzed image only if save_path is provided
    if save_path is not None:
        output_path = os.path.join(save_path, f"{filename}_analyzed.png")
        cv2.imwrite(output_path, image)
        # print(f"Processed image saved: {output_path}")

    return patient_name, all_sections_data

def process_images_in_folder(folder_path, result_folder):
    """Processes all images in the selected folder.""" 
    # Removed creation of result_folder here

    images_processed = 0
    patient_data = {}

    for file in os.listdir(folder_path):
        if file.lower().endswith(('.png', '.jpg', '.jpeg')):
            input_path = os.path.join(folder_path, file)
            base_name, ext = os.path.splitext(file)
            
            # Create patient subfolder under result_folder
            # First get patient name from filename
            filename = os.path.splitext(file)[0]
            parts = filename.split()
            if len(parts) >= 2:
                patient_name = " ".join(parts[:-1]).upper()
            else:
                patient_name = filename.upper()
            
            # Create patient subfolder
            patient_subfolder = os.path.join(result_folder, patient_name)
            os.makedirs(patient_subfolder, exist_ok=True)
            
            # Process image once with the save path
            patient_name, all_sections_data = process_image(input_path, patient_subfolder, None)

            # Accumulate data per patient
            if patient_name not in patient_data:
                patient_data[patient_name] = {
                    "AD": {},
                    "AI": {},
                    "BI": {},
                    "BD": {}
                }
            # Merge all_sections_data into patient_data[patient_name]
            for record_type in all_sections_data:
                for section_index, angles in all_sections_data[record_type].items():
                    patient_data[patient_name][record_type][section_index] = angles

            images_processed += 1

    return patient_data, images_processed

def process_all_subfolders(parent_folder_path):
    """Processes all subfolders in the parent folder and combines patient data."""
    combined_patient_data = {}
    total_images_processed = 0
    result_folder = os.path.join(parent_folder_path, "result")
    os.makedirs(result_folder, exist_ok=True)

    # List all subfolders
    for entry in os.listdir(parent_folder_path):
        subfolder_path = os.path.join(parent_folder_path, entry)
        if os.path.isdir(subfolder_path):
            log_message(f"Processing folder: {subfolder_path}")
            patient_data, images_processed = process_images_in_folder(subfolder_path, result_folder)
            total_images_processed += images_processed

            # Merge patient_data into combined_patient_data
            for patient_name, all_sections_data in patient_data.items():
                if patient_name not in combined_patient_data:
                    combined_patient_data[patient_name] = {
                        "AD": {},
                        "AI": {},
                        "BI": {},
                        "BD": {}
                    }
                for record_type in all_sections_data:
                    for section_index, angles in all_sections_data[record_type].items():
                        combined_patient_data[patient_name][record_type][section_index] = angles

    # Save combined data for all patients
    for patient_name, all_sections_data in combined_patient_data.items():
        save_blue_dots_to_csv(patient_name, all_sections_data, result_folder)

    if total_images_processed > 0:
        log_message(f"✅ Procesamiento completa! {total_images_processed} Procesamiento completa!.")
    else:
        log_message("⚠️ No se encontraron imágenes en las carpetas seleccionadas.")

def select_folder():
    """Opens a dialog to select a parent folder and processes all subfolders."""
    folder_selected = filedialog.askdirectory()
    if folder_selected:
        log_text.delete(1.0, tk.END)  # Clear previous logs
        process_all_subfolders(folder_selected)

def select_single_folder():
    """Opens a dialog to select a folder and processes all images in it."""
    folder_selected = filedialog.askdirectory()
    if folder_selected:
        log_text.delete(1.0, tk.END)  # Clear previous logs
        process_images_in_folder(folder_selected)

def rename_images_based_on_red_curve(parent_folder_path):
    # \"\"\"Renames images in subfolders based on OCR detection of numbers for AXIO images.\"\"\"
    pytesseract.pytesseract.tesseract_cmd = r'C:\\Program Files\\Tesseract-OCR\\tesseract.exe'
    for entry in os.listdir(parent_folder_path):
        subfolder_path = os.path.join(parent_folder_path, entry)
        if os.path.isdir(subfolder_path):
            for file in os.listdir(subfolder_path):
                if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                    input_path = os.path.join(subfolder_path, file)
                    image, gray, edges = preprocess_image(input_path)
                    image_height, image_width = image.shape[:2]

                    ocr_image = cv2.imread(input_path)
                    graph_area = detect_graph_area(edges, gray, image_width)
                    if graph_area is not None:
                        x, y, w, h = graph_area
                        
                        ocr_roi = ocr_image[y:y+h, x:x+112]
                    else:
                        ocr_roi = ocr_image
                    gray_ocr = cv2.cvtColor(ocr_roi, cv2.COLOR_BGR2GRAY)
                    _, thresh_ocr = cv2.threshold(gray_ocr, 150, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
                    custom_config = r'--oem 3 --psm 6 -c tessedit_char_whitelist=0123456789°'
                    text = pytesseract.image_to_string(thresh_ocr, config=custom_config)

                    dir_name = os.path.basename(subfolder_path)
                    base_path = os.path.join(subfolder_path, dir_name)

                    # Check if OCR text contains any digit
                    if any(char.isdigit() for char in text):
                        print(dir_name, text)
                        # Extract numbers from text
                        numbers = [int(num) for num in text.split() if num.isdigit()]
                        if len(numbers) >= 2:
                            # Store axio values for this patient
                            dir_name = os.path.basename(subfolder_path).upper()  # Convert to uppercase to match patient name format
                            patient_axio_values[dir_name] = {
                                'dei': float(numbers[0]),
                                'izq': float(numbers[1])
                            }
                            print(f"Found axio values for {dir_name} - dei: {numbers[0]}, izq: {numbers[1]}")
                            print(f"Current patient_axio_values dictionary: {patient_axio_values}")  # Debug log
                        if not os.path.exists(base_path + " AXIO.png"):
                            new_name = base_path + " AXIO.png"
                        else:
                            new_name_ai = base_path + " AI.png"
                            new_name_bi = base_path + " BI.png"
                            print("Here", new_name_ai, new_name_bi)
                            if not os.path.exists(new_name_ai):
                                new_name = new_name_ai
                            elif not os.path.exists(new_name_bi):
                                new_name = new_name_bi
                            else:
                                # If both AI and BI exist, skip renaming the image
                                new_name = None
                    else:
                        new_name_ai = base_path + " AI.png"
                        new_name_bi = base_path + " BI.png"
                        print("Here", new_name_ai, new_name_bi)
                        if not os.path.exists(new_name_ai):
                            new_name = new_name_ai
                        elif not os.path.exists(new_name_bi):
                            new_name = new_name_bi
                        else:
                            # If both AI and BI exist, skip renaming the image
                            new_name = None
                    # Rename file if different and new_name is not None
                    if new_name and os.path.abspath(input_path) != os.path.abspath(new_name):
                        try:
                            os.rename(input_path, new_name)
                            log_message(f"Renamed {file} to {os.path.basename(new_name)}")
                        except Exception as e:
                            log_message(f"Error renaming {file}: {e}")

def select_parent_folder():
    # \"\"\"Opens a dialog to select a parent folder and processes all subfolders.\"\"\"
    folder_selected = filedialog.askdirectory()
    if folder_selected:
        log_text.delete(1.0, tk.END)  # Clear previous logs
        rename_images_based_on_red_curve(folder_selected)
        process_all_subfolders(folder_selected)

# UI Setup
root = tk.Tk()
root.title("Graph Processing Tool")
root.geometry("500x400")

btn_select_parent_folder = tk.Button(root, text="Multi-Calculate (All Subfolders)", command=select_parent_folder, font=("Arial", 12))
btn_select_parent_folder.pack(pady=10)

log_text = scrolledtext.ScrolledText(root, wrap=tk.WORD, width=60, height=20, font=("Arial", 10))
log_text.pack(pady=10)

root.mainloop()
